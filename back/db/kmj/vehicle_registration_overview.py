from openpyxl import load_workbook
import pandas as pd
from back.db.kmj.db_config import get_conn


def infer_reg_month(filename: str, ws) -> pd.Timestamp:
    """
    기준월을 추정한다.
    1) 파일명에서 YYYY[구분자]MM 패턴 검색
    2) 시트 상단 5행 텍스트에서 YYYY[구분자]MM 검색
    """
    import re
    import pandas as pd

    # 허용 패턴: 2023.01, 2023-01, 2023_01, 202301, 2023년 1월 등
    pattern = re.compile(r"(?P<y>19\d{2}|20\d{2})\D{0,10}(?P<m>1[0-2]|0?[1-9])")

    # 1) 파일명 먼저 시도
    m = pattern.search(filename)
    if m:
        y, mm = int(m.group("y")), int(m.group("m"))
        return pd.Timestamp(y, mm, 1)

    # 2) 워크시트 상단 몇 줄에서 시도 (예: "조회년월: 2023.01")
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        text = " ".join(str(x) for x in row if x is not None)
        m = pattern.search(text)
        if m:
            y, mm = int(m.group("y")), int(m.group("m"))
            return pd.Timestamp(y, mm, 1)

    raise ValueError("기준월(YYYY.MM/년월)을 파일명이나 시트 상단에서 찾지 못했습니다.")


filename = "./data/kmj/2023년_01월_자동차_등록자료_통계.xlsx"

wb = load_workbook(filename=filename, data_only=True)
sheet_name = "04.성별_연령별"
ws = wb[sheet_name]

# (2-1) 헤더 행(2번째 줄) 읽기
header_row_idx = 3  # 1-based (엑셀 줄 번호)
header = [c.value for c in ws[header_row_idx]]


# (2-2) 데이터 행(3번째 줄부터 끝까지) 읽기
data_rows = []
for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
    data_rows.append(list(row))
# print("rows read:", len(data_rows))

# (3-1) DataFrame 생성
df = pd.DataFrame(data_rows, columns=header)

# (3-2) 가짜/빈 컬럼 제거
df = df.loc[:, df.columns.notna()]  # None 컬럼 제거
df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed")]  # Unnamed 제거

# (3-3) 성별 보정(ffill)
if "성별" in df.columns:
    df["성별"] = pd.Series(df["성별"]).ffill().astype(str).str.strip()

    # (3-3-1) 성별 정규화(동의어 통합)
    gender_map = {"남자": "남성", "여자": "여성", "전체": "합계"}
    df["성별"] = df["성별"].replace(gender_map).str.replace(r"\s+", "", regex=True)

    # (3-3-2) 불필요 성별 제거: 합계/계/기타/미상 등
    drop_genders = {"합계", "계", "기타", "미상", "무응답", "불명", ""}
    df = df[~df["성별"].isin(drop_genders)]

# (3-4) 합계 행 제거
if "연령/시도" in df.columns:
    df["연령/시도"] = df["연령/시도"].astype(str).str.strip()
    # 기존: df = df[df["연령/시도"] != "계"]
    drop_age = {"계", "합계", "기타", ""}
    df = df[~df["연령/시도"].isin(drop_age)]


# (3-5) 지역 컬럼 목록 (총계는 DB에 안 넣을 거면 제외)
id_cols = ["성별", "연령/시도"]
region_cols = [c for c in df.columns if c not in id_cols + ["총계"]]

# (3-6) 쉼표 제거 후 숫자화
for c in region_cols:
    df[c] = (
        pd.Series(df[c])
        .astype(str)
        .str.replace(",", "", regex=False)
        .str.replace(" ", "", regex=False)
    )
    df[c] = pd.to_numeric(df[c], errors="coerce")

# === (4) Wide → Long ===
# 안전 체크: 꼭 있어야 할 컬럼
required = set(["성별", "연령/시도"])
missing = required - set(df.columns)
if missing:
    raise KeyError(f"필수 컬럼 누락: {missing}. 실제 컬럼: {list(df.columns)}")

# Long 포맷으로 녹이기
tidy = (
    df.melt(
        id_vars=["성별", "연령/시도"],
        value_vars=region_cols,
        var_name="region",
        value_name="reg_count",
    )
    .dropna(subset=["reg_count"])
    .rename(columns={"성별": "gender", "연령/시도": "age_group"})
)

# 타입/공백 정리
tidy["gender"] = tidy["gender"].astype(str).str.strip()
tidy["age_group"] = tidy["age_group"].astype(str).str.strip()
tidy["region"] = tidy["region"].astype(str).str.strip()
tidy["reg_count"] = tidy["reg_count"].astype("Int64")

reg_month = infer_reg_month(filename, ws)

tidy["reg_month"] = reg_month
tidy = tidy[["reg_month", "region", "gender", "age_group", "reg_count"]]
tidy["reg_month"] = tidy["reg_month"].dt.date  # python date로

tidy = tidy.copy()
tidy["reg_month"] = pd.to_datetime(tidy["reg_month"]).dt.date

insert_sql = """
INSERT INTO vehicle_reg
    (reg_month, region, gender, age_group, reg_count)
VALUES
    (%s, %s, %s, %s, %s)
ON DUPLICATE KEY UPDATE
    reg_count = VALUES(reg_count)
"""

params = [
    (
        row["reg_month"],
        row["region"],
        row["gender"],
        row["age_group"],
        int(row["reg_count"]) if pd.notna(row["reg_count"]) else None,
    )
    for _, row in tidy.iterrows()
]

conn = get_conn()
try:
    with conn.cursor() as cur:
        for i in range(0, len(params), 1000):
            cur.executemany(insert_sql, params[i : i + 1000])
    conn.commit()
finally:
    conn.close()
